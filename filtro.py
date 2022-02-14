from openpyxl import Workbook, load_workbook
from dictionario import dictionary
import time


def initial_excel():
    new_excel = Workbook()
    
    # Add Headers
    new_excel.worksheets[0]['A1'].value = 'ciudad'
    new_excel.worksheets[0]['B1'].value = 'estado'
    new_excel.worksheets[0]['C1'].value = 'clave_Estado'
    new_excel.worksheets[0]['D1'].value = 'lada'
    new_excel.worksheets[0]['E1'].value = 'cifras'
    new_excel.worksheets[0]['F1'].value = 'pais'
    return new_excel


phone_data = load_workbook('telefonia.xlsx')
sheet = phone_data['Hoja1']
columnas = sheet.max_row
contador = 2
new_excel = initial_excel()

# Separate the city from the state key 
def handleCity(fullCity):
    newArray = fullCity.split(',')
    return newArray

# Look for the corresponding state according to its key
def handleSatet(keyState):
    dato  = getState(keyState)
    
    if dato:
        # Si lo encuentra el estado a la primera
        return dato
    else:
        ''' 
            Si no le agregamos el texto extra (Y ÁREA METROPOLITANA) despues de la clave de estado
            Podemos encontrar JAL y JAL Y ÁREA METROPOLITANA 
        '''
        partition = keyState.split(' ')
        dato = getState(partition[0])
        return dato

def getState(KeyState):
    return dictionary.get(KeyState,'')
   
   
while contador <= columnas:
    fullCity = sheet[f'A{contador}'].value
    fullCity = handleCity(fullCity)
    nameState = handleSatet(fullCity[1].strip())
    extPhone = sheet[f'C{contador}'].value.split(' ')[0]
    
    new_excel.worksheets[0][f'A{contador}'].value = fullCity[0]
    new_excel.worksheets[0][f'B{contador}'].value = nameState
    new_excel.worksheets[0][f'C{contador}'].value = fullCity[1]
    new_excel.worksheets[0][f'D{contador}'].value = sheet[f'B{contador}'].value
    new_excel.worksheets[0][f'E{contador}'].value = extPhone
    new_excel.worksheets[0][f'F{contador}'].value = 'Mexico'
    
    contador += 1
    # time.sleep(2)

new_excel.save(f'phone_clean_data.xlsx')